import requests
import time
from openpyxl import load_workbook
import os

CLIENT_ID = "YOUR_CLIENT_ID_HERE"
CLIENT_SECRET = "YOUR_CLIENT_SECRET_HERE"

REGION = "us"
LOCALE = "en_US"
NAMESPACE = f"dynamic-{REGION}"

RANK_MODIFIER_TYPE = 38
RANK_MAP = {
    4: "R1",
    5: "R2",
    6: "R3",
    7: "R4",
    8: "R5",
}

TRACKED_ITEMS = [
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Bands", "item_id": 256880},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Cloak", "item_id": 256891},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Gloves", "item_id": 256887},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Hood", "item_id": 256888},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Leggings", "item_id": 256889},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Sash", "item_id": 256884},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Shoulderpads", "item_id": 256890},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Treads", "item_id": 256886},
    {"group": "Tailoring", "name": "Pattern: Thalassian Competitor's Cloth Tunic", "item_id": 256885},

    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Belt", "item_id": 256631},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Boots", "item_id": 256626},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Chestpiece", "item_id": 256627},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Gloves", "item_id": 256628},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Mask", "item_id": 256632},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Shoulderpads", "item_id": 256630},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Trousers", "item_id": 256629},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Leather Wristwraps", "item_id": 256635},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Cowl", "item_id": 256646},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Cuffs", "item_id": 256654},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Epaulets", "item_id": 256644},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Girdle", "item_id": 256649},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Grips", "item_id": 256643},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Leggings", "item_id": 256641},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Stompers", "item_id": 256633},
    {"group": "Leatherworking", "name": "Pattern: Thalassian Competitor's Chain Tunic", "item_id": 256634},

    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Bow", "item_id": 257258},
    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Insignia of Alacrity", "item_id": 257261},
    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Lamp", "item_id": 257243},
    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Medallion", "item_id": 257262},
    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Pillar", "item_id": 257259},
    {"group": "Inscription", "name": "Technique: Thalassian Competitor's Staff", "item_id": 268366},

    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Bulwark", "item_id": 238229},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Greatsword", "item_id": 238232},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Knife", "item_id": 238227},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Maxim", "item_id": 238228},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Pickaxe", "item_id": 238226},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Armguards", "item_id": 238225},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Breastplate", "item_id": 238218},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Gauntlets", "item_id": 238220},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Greaves", "item_id": 238222},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Helm", "item_id": 238221},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Pauldrons", "item_id": 238223},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Sabatons", "item_id": 238219},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Plate Waistguard", "item_id": 238224},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Skewer", "item_id": 238231},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Splitter", "item_id": 238230},
    {"group": "Blacksmithing", "name": "Plans: Thalassian Competitor's Sword", "item_id": 238233},

    {"group": "Jewelcrafting", "name": "Design: Thalassian Competitor's Amulet", "item_id": 256719},
    {"group": "Jewelcrafting", "name": "Design: Thalassian Competitor's Signet", "item_id": 256702},
]

OUTPUT_ORDER = [
    "Leatherworking",
    "Inscription",
    "Jewelcrafting",
    "Blacksmithing",
    "Tailoring",
]

def get_access_token():
    url = f"https://{REGION}.battle.net/oauth/token"
    response = requests.post(
        url,
        data={"grant_type": "client_credentials"},
        auth=(CLIENT_ID, CLIENT_SECRET),
        timeout=30
    )
    response.raise_for_status()
    return response.json()["access_token"]


def api_get(path, token, max_retries=3):
    url = f"https://{REGION}.api.blizzard.com{path}"

    params = {
        "namespace": NAMESPACE,
        "locale": LOCALE
    }

    headers = {
        "Authorization": f"Bearer {token}"
    }

    for attempt in range(1, max_retries + 1):
        response = requests.get(url, params=params, headers=headers, timeout=60)

        if response.status_code in (502, 503, 504):
            print(f"Retryable error {response.status_code} for {path} (attempt {attempt}/{max_retries})")
            if attempt < max_retries:
                time.sleep(attempt * 2)
                continue

        response.raise_for_status()
        return response.json()


def get_connected_realm_index(token):
    return api_get("/data/wow/connected-realm/index", token)


def get_connected_realm_detail(token, connected_realm_id):
    return api_get(f"/data/wow/connected-realm/{connected_realm_id}", token)


def get_connected_realm_auctions(token, connected_realm_id):
    return api_get(f"/data/wow/connected-realm/{connected_realm_id}/auctions", token)


def extract_connected_realm_id_from_href(href):
    tail = href.rstrip("/").split("/")[-1]
    return int(tail.split("?")[0])


def build_connected_realm_lookup(token):
    index_data = get_connected_realm_index(token)
    connected_realms = index_data.get("connected_realms", [])

    connected_lookup = {}

    for entry in connected_realms:
        href = entry["href"]
        connected_id = extract_connected_realm_id_from_href(href)

        detail = get_connected_realm_detail(token, connected_id)
        realm_names = []

        for realm in detail.get("realms", []):
            realm_name = realm.get("name")
            if realm_name:
                realm_names.append(realm_name)

        connected_lookup[connected_id] = realm_names

    return connected_lookup

def format_price(copper):
    gold = copper // 10000
    silver = (copper % 10000) // 100
    copper_only = copper % 100

    parts = []

    if gold > 0:
        parts.append(f"{gold}g")
    if silver > 0:
        parts.append(f"{silver}s")
    if copper_only > 0 or not parts:
        parts.append(f"{copper_only}c")

    return " ".join(parts)


def scan_all_tracked_items_across_all_realms(token, connected_lookup, tracked_items):
    tracked_item_ids = {item["item_id"] for item in tracked_items}
    all_results = {item_id: [] for item_id in tracked_item_ids}

    total = len(connected_lookup)

    for index, (connected_id, realm_names) in enumerate(connected_lookup.items(), start=1):
        print(f"Scanning {index}/{total}: connected realm {connected_id}")

        try:
            auction_data = get_connected_realm_auctions(token, connected_id)
        except requests.HTTPError as ex:
            print(f"Skipping connected realm {connected_id} due to error: {ex}")
            continue

        for auction in auction_data.get("auctions", []):
            item = auction.get("item", {})
            item_id = item.get("id")

            if item_id not in tracked_item_ids:
                continue

            buyout = auction.get("buyout")
            if buyout is None:
                continue

            all_results[item_id].append({
                "connected_realm_id": connected_id,
                "realms": realm_names,
                "auction_id": auction.get("id"),
                "item_id": item_id,
                "buyout": buyout,
                "quantity": auction.get("quantity"),
                "time_left": auction.get("time_left")
            })

    for item_id in all_results:
        all_results[item_id].sort(key=lambda x: x["buyout"])

    return all_results


def get_cheapest_result(all_matches):
    if not all_matches:
        return None
    return all_matches[0]


def scan_all_tracked_items_across_all_realms(token, connected_lookup, tracked_items):
    tracked_item_ids = {item["item_id"] for item in tracked_items}
    all_results = {item_id: [] for item_id in tracked_item_ids}

    total = len(connected_lookup)

    for index, (connected_id, realm_names) in enumerate(connected_lookup.items(), start=1):
        print(f"Scanning {index}/{total}: connected realm {connected_id}")

        try:
            auction_data = get_connected_realm_auctions(token, connected_id)
        except requests.HTTPError as ex:
            print(f"Skipping connected realm {connected_id} due to error: {ex}")
            continue

        for auction in auction_data.get("auctions", []):
            item = auction.get("item", {})
            item_id = item.get("id")

            if item_id not in tracked_item_ids:
                continue

            buyout = auction.get("buyout")
            if buyout is None:
                continue

            all_results[item_id].append({
                "connected_realm_id": connected_id,
                "realms": realm_names,
                "auction_id": auction.get("id"),
                "item_id": item_id,
                "item_data": item,
                "buyout": buyout,
                "quantity": auction.get("quantity"),
                "time_left": auction.get("time_left")
            })

    for item_id in all_results:
        all_results[item_id].sort(key=lambda x: x["buyout"])

    return all_results


def build_result_row(item_group, item_name, item_id, all_matches):
    cheapest = get_cheapest_result(all_matches)

    row = {
        "group": item_group,
        "item_name": item_name,
        "item_id": item_id,
        "cheapest_price": "",
        "cheapest_realms": "",
        "total_matches": len(all_matches),
    }

    if cheapest:
        row["cheapest_price"] = format_price(cheapest["buyout"])
        row["cheapest_realms"] = ", ".join(cheapest["realms"])

    return row


def write_results_to_excel(result_rows):
    path = r"C:\Users\codyw\Desktop\AHScanResults.xlsx"

    wb = load_workbook(path)
    ws = wb["Recipe"]

    # clear old results but keep header row
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in result_rows:
        ws.append([
            row["group"],
            row["item_name"],
            row["item_id"],
            row["cheapest_price"],
            row["cheapest_realms"],
            row["total_matches"],
        ])

    wb.save(path)
    print("\nResults written to AHScanResults.xlsx")
    os.startfile(path)


def main():
    token = get_access_token()

    connected_lookup = build_connected_realm_lookup(token)
    all_results = scan_all_tracked_items_across_all_realms(token, connected_lookup, TRACKED_ITEMS)

    result_rows = []

    for item in TRACKED_ITEMS:
        item_id = item["item_id"]
        item_name = item["name"]
        item_group = item["group"]
        all_matches = all_results.get(item_id, [])

        row = build_result_row(item_group, item_name, item_id, all_matches)
        result_rows.append(row)

    write_results_to_excel(result_rows)


if __name__ == "__main__":
    main()