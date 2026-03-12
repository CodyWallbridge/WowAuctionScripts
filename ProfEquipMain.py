import requests
import time
from openpyxl import load_workbook
import os

CLIENT_ID = "YOUR_CLIENT_ID_HERE"
CLIENT_SECRET = "YOUR_CLIENT_SECRET_HERE"

REGION = "us"
LOCALE = "en_US"
NAMESPACE = f"dynamic-{REGION}"

RANK_MODIFIER_TYPE = 38
RANK_MAP = {
    4: "R1",
    5: "R2",
    6: "R3",
    7: "R4",
    8: "R5",
}

TRACKED_ITEMS = [
    {"name": "Sin'dorei Alchemist's Mixing Rod", "item_id": 245778},
    {"name": "Sin'dorei Alchemist's Hat", "item_id": 244626},
    {"name": "Elegant Artisan's Alchemy Coveralls", "item_id": 239635},

    {"name": "Sun-Blessed Blacksmith's Hammer", "item_id": 238018},
    {"name": "Sun-Blessed Blacksmith's Toolbox", "item_id": 237952},
    {"name": "Sin'dorei Forgemaster's Cover", "item_id": 244628},

    {"name": "Sin'dorei Leathershaper's Smock", "item_id": 244625},
    {"name": "Sun-Blessed Leatherworker's Toolset", "item_id": 237951},
    {"name": "Sun-Blessed Leatherworker's Knife", "item_id": 238017},

    {"name": "Runed Brilliant Silver Rod", "item_id": 244176},
    {"name": "Elegant Artisan's Enchanting Hat", "item_id": 239637},
    {"name": "Sin'dorei Enchanter's Crystal", "item_id": 240960},

    {"name": "Sin'dorei Snippers", "item_id": 244708},
    {"name": "Sun-Blessed Needle Set", "item_id": 237950},
    {"name": "Elegant Artisan's Tailoring Robe", "item_id": 239640},

    {"name": "Sin'dorei Quill", "item_id": 245776},
    {"name": "Sin'dorei Scribe's Spectacles", "item_id": 240957},
    {"name": "Improved Right-Handed Magnifying Glass", "item_id": 240958},

    {"name": "Sin'dorei Clampers", "item_id": 244714},
    {"name": "Sin'dorei Jeweler's Loupes", "item_id": 240959},
    {"name": "Sin'dorei Jeweler's Cover", "item_id": 244630},
]

def get_access_token():
    url = f"https://{REGION}.battle.net/oauth/token"
    response = requests.post(
        url,
        data={"grant_type": "client_credentials"},
        auth=(CLIENT_ID, CLIENT_SECRET),
        timeout=30
    )
    response.raise_for_status()
    return response.json()["access_token"]


def api_get(path, token, max_retries=3):
    url = f"https://{REGION}.api.blizzard.com{path}"

    params = {
        "namespace": NAMESPACE,
        "locale": LOCALE
    }

    headers = {
        "Authorization": f"Bearer {token}"
    }

    for attempt in range(1, max_retries + 1):
        response = requests.get(url, params=params, headers=headers, timeout=60)

        if response.status_code in (502, 503, 504):
            print(f"Retryable error {response.status_code} for {path} (attempt {attempt}/{max_retries})")
            if attempt < max_retries:
                time.sleep(attempt * 2)
                continue

        response.raise_for_status()
        return response.json()


def get_connected_realm_index(token):
    return api_get("/data/wow/connected-realm/index", token)


def get_connected_realm_detail(token, connected_realm_id):
    return api_get(f"/data/wow/connected-realm/{connected_realm_id}", token)


def get_connected_realm_auctions(token, connected_realm_id):
    return api_get(f"/data/wow/connected-realm/{connected_realm_id}/auctions", token)


def extract_connected_realm_id_from_href(href):
    tail = href.rstrip("/").split("/")[-1]
    return int(tail.split("?")[0])


def build_connected_realm_lookup(token):
    index_data = get_connected_realm_index(token)
    connected_realms = index_data.get("connected_realms", [])

    connected_lookup = {}

    for entry in connected_realms:
        href = entry["href"]
        connected_id = extract_connected_realm_id_from_href(href)

        detail = get_connected_realm_detail(token, connected_id)
        realm_names = []

        for realm in detail.get("realms", []):
            realm_name = realm.get("name")
            if realm_name:
                realm_names.append(realm_name)

        connected_lookup[connected_id] = realm_names

    return connected_lookup


def find_auctions_for_item(auction_data, item_id):
    matches = []

    for auction in auction_data.get("auctions", []):
        item = auction.get("item", {})
        auction_item_id = item.get("id")

        if auction_item_id != item_id:
            continue

        buyout = auction.get("buyout")
        if buyout is None:
            continue

        matches.append({
            "auction_id": auction.get("id"),
            "item_id": auction_item_id,
            "item_data": item,
            "buyout": buyout,
            "quantity": auction.get("quantity"),
            "time_left": auction.get("time_left")
        })

    matches.sort(key=lambda x: x["buyout"])
    return matches


def scan_item_across_all_realms(token, connected_lookup, item_id):
    all_matches = []

    total = len(connected_lookup)

    for index, (connected_id, realm_names) in enumerate(connected_lookup.items(), start=1):
        print(f"Scanning {index}/{total}: connected realm {connected_id}")

        try:
            auction_data = get_connected_realm_auctions(token, connected_id)
        except requests.HTTPError as ex:
            print(f"Skipping connected realm {connected_id} due to error: {ex}")
            continue

        matches = find_auctions_for_item(auction_data, item_id)

        for match in matches:
            all_matches.append({
                "connected_realm_id": connected_id,
                "realms": realm_names,
                "auction_id": match["auction_id"],
                "item_id": match["item_id"],
                "item_data": match["item_data"],
                "buyout": match["buyout"],
                "quantity": match["quantity"],
                "time_left": match["time_left"]
            })

    all_matches.sort(key=lambda x: x["buyout"])
    return all_matches


def format_price(copper):
    gold = copper // 10000
    silver = (copper % 10000) // 100
    copper_only = copper % 100

    parts = []

    if gold > 0:
        parts.append(f"{gold}g")
    if silver > 0:
        parts.append(f"{silver}s")
    if copper_only > 0 or not parts:
        parts.append(f"{copper_only}c")

    return " ".join(parts)


def get_modifier_value(item_data, modifier_type):
    for mod in item_data.get("modifiers", []):
        if mod.get("type") == modifier_type:
            return mod.get("value")
    return None


def get_rank_label(item_data):
    modifier_value = get_modifier_value(item_data, RANK_MODIFIER_TYPE)
    return RANK_MAP.get(modifier_value, f"Unknown ({modifier_value})")


def get_lowest_unique_realm_results(all_matches, top_n=3):
    results = []
    seen_connected_realms = set()

    for match in all_matches:
        connected_id = match["connected_realm_id"]

        if connected_id in seen_connected_realms:
            continue

        seen_connected_realms.add(connected_id)
        results.append(match)

        if len(results) == top_n:
            break

    return results


def get_cheapest_rank_result(all_matches, rank_label):
    for match in all_matches:
        if get_rank_label(match["item_data"]) == rank_label:
            return match
    return None


def summarize_item_results(item_id, all_matches, top_n=3):
    lowest_unique = get_lowest_unique_realm_results(all_matches, top_n)
    cheapest_r5 = get_cheapest_rank_result(all_matches, "R5")

    summary = {
        "item_id": item_id,
        "total_matches": len(all_matches),
        "top_results": [],
        "cheapest_r5": None
    }

    for match in lowest_unique:
        summary["top_results"].append({
            "buyout": match["buyout"],
            "rank": get_rank_label(match["item_data"]),
            "realms": ", ".join(match["realms"])
        })

    if cheapest_r5:
        summary["cheapest_r5"] = {
            "buyout": cheapest_r5["buyout"],
            "rank": get_rank_label(cheapest_r5["item_data"]),
            "realms": ", ".join(cheapest_r5["realms"])
        }

    return summary


def scan_all_tracked_items_across_all_realms(token, connected_lookup, tracked_items):
    tracked_item_ids = {item["item_id"] for item in tracked_items}
    all_results = {item_id: [] for item_id in tracked_item_ids}

    total = len(connected_lookup)

    for index, (connected_id, realm_names) in enumerate(connected_lookup.items(), start=1):
        print(f"Scanning {index}/{total}: connected realm {connected_id}")

        try:
            auction_data = get_connected_realm_auctions(token, connected_id)
        except requests.HTTPError as ex:
            print(f"Skipping connected realm {connected_id} due to error: {ex}")
            continue

        for auction in auction_data.get("auctions", []):
            item = auction.get("item", {})
            item_id = item.get("id")

            if item_id not in tracked_item_ids:
                continue

            buyout = auction.get("buyout")
            if buyout is None:
                continue

            all_results[item_id].append({
                "connected_realm_id": connected_id,
                "realms": realm_names,
                "auction_id": auction.get("id"),
                "item_id": item_id,
                "item_data": item,
                "buyout": buyout,
                "quantity": auction.get("quantity"),
                "time_left": auction.get("time_left")
            })

    for item_id in all_results:
        all_results[item_id].sort(key=lambda x: x["buyout"])

    return all_results


def build_result_row(item_name, item_id, all_matches):
    summary = summarize_item_results(item_id, all_matches, top_n=3)

    row = {
        "item_name": item_name,
        "item_id": item_id,
        "total_matches": summary["total_matches"],

        "lowest_1_price": "",
        "lowest_1_rank": "",
        "lowest_1_realms": "",

        "lowest_2_price": "",
        "lowest_2_rank": "",
        "lowest_2_realms": "",

        "lowest_3_price": "",
        "lowest_3_rank": "",
        "lowest_3_realms": "",

        "cheapest_r5_price": "",
        "cheapest_r5_rank": "",
        "cheapest_r5_realms": "",
    }

    for i, result in enumerate(summary["top_results"], start=1):
        row[f"lowest_{i}_price"] = format_price(result["buyout"])
        row[f"lowest_{i}_rank"] = result["rank"]
        row[f"lowest_{i}_realms"] = result["realms"]

    if summary["cheapest_r5"]:
        row["cheapest_r5_price"] = format_price(summary["cheapest_r5"]["buyout"])
        row["cheapest_r5_rank"] = summary["cheapest_r5"]["rank"]
        row["cheapest_r5_realms"] = summary["cheapest_r5"]["realms"]

    return row


def write_results_to_excel(result_rows):
    path = r"C:\Users\codyw\Desktop\AHScanResults.xlsx"
    
    wb = load_workbook(path)
    ws = wb["Prof Equip"]

    # clear old results (leave header)
    ws.delete_rows(2, ws.max_row)

    for row in result_rows:
        ws.append([
            row["item_name"],
            row["item_id"],
            row["total_matches"],

            row["lowest_1_price"],
            row["lowest_1_rank"],
            row["lowest_1_realms"],

            row["lowest_2_price"],
            row["lowest_2_rank"],
            row["lowest_2_realms"],

            row["lowest_3_price"],
            row["lowest_3_rank"],
            row["lowest_3_realms"],

            row["cheapest_r5_price"],
            row["cheapest_r5_rank"],
            row["cheapest_r5_realms"]
        ])

    wb.save(path)
    print("\nResults written to AHScanResults.xlsx")
    os.startfile(path)


def main():
    token = get_access_token()

    connected_lookup = build_connected_realm_lookup(token)
    all_results = scan_all_tracked_items_across_all_realms(token, connected_lookup, TRACKED_ITEMS)

    result_rows = []

    for item in TRACKED_ITEMS:
        item_id = item["item_id"]
        item_name = item["name"]
        all_matches = all_results.get(item_id, [])

        row = build_result_row(item_name, item_id, all_matches)
        result_rows.append(row)

    write_results_to_excel(result_rows)


if __name__ == "__main__":
    main()